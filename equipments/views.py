from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

from .forms import (
    EchographeForm, EtablissementForm, FournisseurForm, MaintenanceForm,
    MouvementStockForm, PanneForm, PieceRechangeForm,
    PlanificationMaintenanceForm, ServiceForm,
)
from .models import (
    Echographe, Etablissement, Fournisseur, Maintenance, MouvementStock,
    Notification, Panne, PieceRechange, PlanificationMaintenance, Service,
    Suggestion,
)
from .utils import (
    generer_qr_code, marquer_suggestion_traitee, notifier_ajout,
    suggerer_maintenance_echographe, suggerer_panne_recurrente,
    suggerer_stock_bas, suggerer_technicien_frequent,
)

# ==========================================================================
#  RÔLES & PERMISSIONS
#  Groupes : "Responsable" (tout), "Technicien" (ajout/modification),
#  "Lecture seule" (consultation uniquement)
# ==========================================================================

GROUPE_RESPONSABLE = "Responsable"
GROUPE_TECHNICIEN = "Technicien"
GROUPE_LECTURE_SEULE = "Lecture seule"


def user_peut_modifier(user):
    """Responsable et Technicien peuvent ajouter/modifier. Lecture seule ne peut pas.
    Les superutilisateurs et le staff ont toujours accès."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    groupes = set(user.groups.values_list('name', flat=True))
    if GROUPE_LECTURE_SEULE in groupes and not (groupes & {GROUPE_RESPONSABLE, GROUPE_TECHNICIEN}):
        return False
    if not groupes:
        # Utilisateur sans groupe assigné : accès en modification par défaut
        return True
    return bool(groupes & {GROUPE_RESPONSABLE, GROUPE_TECHNICIEN})


def user_peut_supprimer(user):
    """Seul le rôle Responsable (ou le staff) peut supprimer."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    groupes = set(user.groups.values_list('name', flat=True))
    if not groupes:
        return True
    return GROUPE_RESPONSABLE in groupes


class ModificationRequiredMixin(LoginRequiredMixin):
    """Bloque l'accès aux vues de création/modification pour le rôle Lecture seule."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and not user_peut_modifier(request.user):
            messages.error(request, "Votre rôle (Lecture seule) ne permet pas cette action.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        return super().dispatch(request, *args, **kwargs)


class SuppressionRequiredMixin(LoginRequiredMixin):
    """Réserve la suppression au rôle Responsable."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and not user_peut_supprimer(request.user):
            messages.error(request, "Seul le rôle Responsable peut supprimer un élément.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        return super().dispatch(request, *args, **kwargs)


# ==========================================================================
#  TABLEAU DE BORD
# ==========================================================================

class DashboardView(LoginRequiredMixin, ListView):
    template_name = 'dashboard.html'
    context_object_name = 'notifications_feed'
    model = Notification
    paginate_by = 0

    def get_queryset(self):
        return Notification.objects.select_related('auteur')[:10]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        aujourdhui = timezone.now().date()
        horizon = aujourdhui + timedelta(days=30)

        ctx['total_echographes'] = Echographe.objects.count()
        ctx['total_services'] = Service.objects.count()
        ctx['total_etablissements'] = Etablissement.objects.count()
        ctx['total_membres'] = User.objects.filter(is_active=True).count()

        ctx['pannes_en_cours'] = Panne.objects.exclude(statut='resolue').count()
        ctx['pannes_urgentes'] = Panne.objects.filter(
            priorite='urgente'
        ).exclude(statut='resolue').count()

        ctx['maintenances_a_venir'] = PlanificationMaintenance.objects.filter(
            date_prevue__range=[aujourdhui, horizon]
        ).exclude(statut__in=['terminee', 'annulee']).order_by('date_prevue')[:6]

        ctx['pieces_en_alerte'] = [
            p for p in PieceRechange.objects.select_related('fournisseur')
            if p.en_alerte
        ]

        ctx['dernieres_pannes'] = Panne.objects.select_related('equipement')[:5]
        ctx['derniers_echographes'] = Echographe.objects.select_related('service')[:5]

        # Répartition des pannes par statut (pour le graphique)
        ctx['pannes_par_statut'] = {
            'declaree': Panne.objects.filter(statut='declaree').count(),
            'en_cours': Panne.objects.filter(statut='en_cours').count(),
            'resolue': Panne.objects.filter(statut='resolue').count(),
        }

        ctx['suggestions_actives'] = Suggestion.objects.filter(traitee=False)[:6]
        ctx['suggestions_actives_count'] = Suggestion.objects.filter(traitee=False).count()
        return ctx


# ==========================================================================
#  NOTIFICATIONS
# ==========================================================================

class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'equipments/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.select_related('auteur')


def marquer_notification_lue(request, pk):
    notification = get_object_or_404(Notification, pk=pk)
    notification.lu_par.add(request.user)
    if notification.lien:
        return redirect(notification.lien)
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


def marquer_toutes_lues(request):
    for notification in Notification.objects.exclude(lu_par=request.user):
        notification.lu_par.add(request.user)
    messages.success(request, "Toutes les notifications ont été marquées comme lues.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ==========================================================================
#  SUGGESTIONS AUTOMATIQUES
# ==========================================================================

class SuggestionListView(LoginRequiredMixin, ListView):
    model = Suggestion
    template_name = 'equipments/suggestion_list.html'
    context_object_name = 'suggestions'
    paginate_by = 20

    def get_queryset(self):
        return Suggestion.objects.filter(traitee=False)


def marquer_suggestion_vue(request, pk):
    suggestion = get_object_or_404(Suggestion, pk=pk)
    marquer_suggestion_traitee(suggestion, request.user)
    messages.success(request, "Suggestion marquée comme traitée.")
    return redirect(request.META.get('HTTP_REFERER', 'suggestion_list'))


# ==========================================================================
#  MIXIN : notifie tous les membres à la création d'un nouvel objet
# ==========================================================================

class NotifiantCreateView(ModificationRequiredMixin, CreateView):
    notif_type = 'info'
    notif_titre = "Nouvel élément ajouté"

    def message_notification(self, objet):
        return f"{self.request.user.get_full_name() or self.request.user.username} a ajouté : {objet}"

    def generer_suggestions(self, objet):
        """Point d'extension : les vues concernées surchargent cette méthode
        pour déclencher le moteur de suggestions après l'enregistrement."""
        return None

    def form_valid(self, form):
        response = super().form_valid(form)
        notifier_ajout(
            auteur=self.request.user,
            type_notification=self.notif_type,
            titre=self.notif_titre,
            message=self.message_notification(self.object),
            lien=self.get_success_url(),
        )
        self.generer_suggestions(self.object)
        messages.success(self.request, "Ajout effectué avec succès. Les membres ont été notifiés.")
        return response


# ==========================================================================
#  ÉTABLISSEMENTS
# ==========================================================================

class EtablissementListView(LoginRequiredMixin, ListView):
    model = Etablissement
    template_name = 'equipments/etablissement_list.html'
    context_object_name = 'objets'


class EtablissementCreateView(NotifiantCreateView):
    model = Etablissement
    form_class = EtablissementForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('etablissement_list')
    notif_type = 'info'
    notif_titre = "Nouvel établissement"
    extra_context = {'titre_page': "Ajouter un établissement"}


class EtablissementUpdateView(ModificationRequiredMixin, UpdateView):
    model = Etablissement
    form_class = EtablissementForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('etablissement_list')
    extra_context = {'titre_page': "Modifier l'établissement"}


class EtablissementDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Etablissement
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('etablissement_list')


# ==========================================================================
#  SERVICES
# ==========================================================================

class ServiceListView(LoginRequiredMixin, ListView):
    model = Service
    template_name = 'equipments/service_list.html'
    context_object_name = 'objets'


class ServiceCreateView(NotifiantCreateView):
    model = Service
    form_class = ServiceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('service_list')
    notif_type = 'info'
    notif_titre = "Nouveau service"
    extra_context = {'titre_page': "Ajouter un service"}


class ServiceUpdateView(ModificationRequiredMixin, UpdateView):
    model = Service
    form_class = ServiceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('service_list')
    extra_context = {'titre_page': "Modifier le service"}


class ServiceDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Service
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('service_list')


# ==========================================================================
#  FOURNISSEURS
# ==========================================================================

class FournisseurListView(LoginRequiredMixin, ListView):
    model = Fournisseur
    template_name = 'equipments/fournisseur_list.html'
    context_object_name = 'objets'


class FournisseurCreateView(NotifiantCreateView):
    model = Fournisseur
    form_class = FournisseurForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('fournisseur_list')
    notif_type = 'info'
    notif_titre = "Nouveau fournisseur"
    extra_context = {'titre_page': "Ajouter un fournisseur"}


class FournisseurUpdateView(ModificationRequiredMixin, UpdateView):
    model = Fournisseur
    form_class = FournisseurForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('fournisseur_list')
    extra_context = {'titre_page': "Modifier le fournisseur"}


class FournisseurDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Fournisseur
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('fournisseur_list')


# ==========================================================================
#  ÉCHOGRAPHES
# ==========================================================================

class EchographeListView(LoginRequiredMixin, ListView):
    model = Echographe
    template_name = 'equipments/echographe_list.html'
    context_object_name = 'objets'
    paginate_by = 12

    def get_queryset(self):
        qs = Echographe.objects.select_related('service', 'service__etablissement')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(numero_serie__icontains=q) | qs.filter(marque__icontains=q) | qs.filter(modele__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q', '')
        return ctx


class EchographeDetailView(LoginRequiredMixin, DetailView):
    model = Echographe
    template_name = 'equipments/echographe_detail.html'
    context_object_name = 'objet'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['maintenances'] = self.object.maintenances.all()
        ctx['planifications'] = self.object.planifications.all()
        ctx['pannes'] = self.object.pannes.all()
        return ctx


class EchographeCreateView(NotifiantCreateView):
    model = Echographe
    form_class = EchographeForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('echographe_list')
    notif_type = 'echographe'
    notif_titre = "Nouvel échographe ajouté"
    extra_context = {'titre_page': "Ajouter un échographe"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return (
            f"{auteur} a ajouté un nouvel échographe : {objet.marque} {objet.modele} "
            f"(N° série : {objet.numero_serie}) au service {objet.service}."
        )

    def form_valid(self, form):
        response = super().form_valid(form)
        # Génération du QR code après l'enregistrement (nécessite `request`
        # pour construire l'URL absolue, indisponible dans un signal).
        generer_qr_code(self.object, self.request)
        return response

    def generer_suggestions(self, objet):
        suggerer_maintenance_echographe(objet)


class EchographeUpdateView(ModificationRequiredMixin, UpdateView):
    model = Echographe
    form_class = EchographeForm
    template_name = 'equipments/generic_form.html'
    extra_context = {'titre_page': "Modifier l'échographe"}

    def get_success_url(self):
        return reverse_lazy('echographe_detail', args=[self.object.pk])

    def form_valid(self, form):
        response = super().form_valid(form)
        if not self.object.qr_code:
            generer_qr_code(self.object, self.request)
        return response


class EchographeDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Echographe
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('echographe_list')


# ==========================================================================
#  QR CODE — SCAN
# ==========================================================================

def qr_scan_view(request):
    """
    Page de scan : l'utilisateur saisit ou scanne le contenu d'un QR code
    (URL de la fiche ou n° de série) et est redirigé vers l'équipement.
    """
    from .models import QRImport

    if request.method == 'POST':
        contenu = request.POST.get('contenu', '').strip()
        QRImport.objects.create(contenu_qr=contenu)

        echographe = None
        # Cas 1 : le contenu est l'URL absolue de la fiche (ex : .../echographes/12/)
        segments = [s for s in contenu.split('/') if s]
        if segments and segments[-1].isdigit():
            echographe = Echographe.objects.filter(pk=segments[-1]).first()

        # Cas 2 : le contenu est directement le n° de série
        if echographe is None:
            echographe = Echographe.objects.filter(numero_serie=contenu).first()

        if echographe:
            messages.success(request, f"Équipement trouvé : {echographe}")
            return redirect('echographe_detail', pk=echographe.pk)

        messages.error(request, "Aucun équipement ne correspond à ce QR code.")
        return redirect('qr_scan')

    return render(request, 'equipments/qr_scan.html')


# ==========================================================================
#  MAINTENANCES
# ==========================================================================

class MaintenanceListView(LoginRequiredMixin, ListView):
    model = Maintenance
    template_name = 'equipments/maintenance_list.html'
    context_object_name = 'objets'
    paginate_by = 15

    def get_queryset(self):
        return Maintenance.objects.select_related('echographe')


class MaintenanceCreateView(NotifiantCreateView):
    model = Maintenance
    form_class = MaintenanceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('maintenance_list')
    notif_type = 'maintenance'
    notif_titre = "Nouvelle intervention de maintenance"
    extra_context = {'titre_page': "Ajouter une maintenance"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return (
            f"{auteur} a enregistré une maintenance {objet.get_type_maintenance_display().lower()} "
            f"sur {objet.echographe} (technicien : {objet.technicien})."
        )

    def generer_suggestions(self, objet):
        suggerer_maintenance_echographe(objet.echographe)


class MaintenanceUpdateView(ModificationRequiredMixin, UpdateView):
    model = Maintenance
    form_class = MaintenanceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('maintenance_list')
    extra_context = {'titre_page': "Modifier la maintenance"}


class MaintenanceDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Maintenance
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('maintenance_list')


# ==========================================================================
#  PLANIFICATIONS DE MAINTENANCE
# ==========================================================================

class PlanificationListView(LoginRequiredMixin, ListView):
    model = PlanificationMaintenance
    template_name = 'equipments/planification_list.html'
    context_object_name = 'objets'

    def get_queryset(self):
        return PlanificationMaintenance.objects.select_related('echographe')


class PlanificationCreateView(NotifiantCreateView):
    model = PlanificationMaintenance
    form_class = PlanificationMaintenanceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('planification_list')
    notif_type = 'planification'
    notif_titre = "Nouvelle planification de maintenance"
    extra_context = {'titre_page': "Planifier une maintenance"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return (
            f"{auteur} a planifié une maintenance {objet.get_type_maintenance_display().lower()} "
            f"pour {objet.echographe} le {objet.date_prevue:%d/%m/%Y}."
        )


class PlanificationUpdateView(ModificationRequiredMixin, UpdateView):
    model = PlanificationMaintenance
    form_class = PlanificationMaintenanceForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('planification_list')
    extra_context = {'titre_page': "Modifier la planification"}


class PlanificationDeleteView(SuppressionRequiredMixin, DeleteView):
    model = PlanificationMaintenance
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('planification_list')


# ==========================================================================
#  PANNES
# ==========================================================================

class PanneListView(LoginRequiredMixin, ListView):
    model = Panne
    template_name = 'equipments/panne_list.html'
    context_object_name = 'objets'
    paginate_by = 15

    def get_queryset(self):
        return Panne.objects.select_related('equipement', 'technicien')


class PanneCreateView(NotifiantCreateView):
    model = Panne
    form_class = PanneForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('panne_list')
    notif_type = 'panne'
    notif_titre = "Nouvelle panne déclarée"
    extra_context = {'titre_page': "Déclarer une panne"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return (
            f"{auteur} a déclaré une panne (priorité {objet.get_priorite_display().lower()}) "
            f"sur {objet.equipement}."
        )

    def generer_suggestions(self, objet):
        suggerer_panne_recurrente(objet)
        suggerer_technicien_frequent(objet)


class PanneUpdateView(ModificationRequiredMixin, UpdateView):
    model = Panne
    form_class = PanneForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('panne_list')
    extra_context = {'titre_page': "Modifier la panne"}


class PanneDeleteView(SuppressionRequiredMixin, DeleteView):
    model = Panne
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('panne_list')


# ==========================================================================
#  PIÈCES DE RECHANGE / STOCK
# ==========================================================================

class PieceRechangeListView(LoginRequiredMixin, ListView):
    model = PieceRechange
    template_name = 'equipments/piece_list.html'
    context_object_name = 'objets'

    def get_queryset(self):
        return PieceRechange.objects.select_related('fournisseur')


class PieceRechangeCreateView(NotifiantCreateView):
    model = PieceRechange
    form_class = PieceRechangeForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('piece_list')
    notif_type = 'piece'
    notif_titre = "Nouvelle pièce de rechange"
    extra_context = {'titre_page': "Ajouter une pièce de rechange"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return f"{auteur} a ajouté la pièce de rechange « {objet.nom} » (réf. {objet.reference}) au stock."

    def generer_suggestions(self, objet):
        suggerer_stock_bas(objet)


class PieceRechangeUpdateView(ModificationRequiredMixin, UpdateView):
    model = PieceRechange
    form_class = PieceRechangeForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('piece_list')
    extra_context = {'titre_page': "Modifier la pièce de rechange"}

    def form_valid(self, form):
        response = super().form_valid(form)
        suggerer_stock_bas(self.object)
        return response


class PieceRechangeDeleteView(SuppressionRequiredMixin, DeleteView):
    model = PieceRechange
    template_name = 'equipments/generic_confirm_delete.html'
    success_url = reverse_lazy('piece_list')


class MouvementStockCreateView(NotifiantCreateView):
    model = MouvementStock
    form_class = MouvementStockForm
    template_name = 'equipments/generic_form.html'
    success_url = reverse_lazy('piece_list')
    notif_type = 'stock'
    notif_titre = "Nouveau mouvement de stock"
    extra_context = {'titre_page': "Enregistrer un mouvement de stock"}

    def message_notification(self, objet):
        auteur = self.request.user.get_full_name() or self.request.user.username
        return (
            f"{auteur} a enregistré une {objet.get_type_mouvement_display().lower()} de "
            f"{objet.quantite} unité(s) pour la pièce « {objet.piece.nom} »."
        )

    def generer_suggestions(self, objet):
        suggerer_stock_bas(objet.piece)


# ==========================================================================
#  EXPORTS PDF / EXCEL
# ==========================================================================

def export_maintenances_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="maintenances.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Rapport des maintenances — GMAO Biomédicale", styles['Title'])]

    data = [["Échographe", "Type", "Date", "Technicien", "Panne", "Réparation", "Temps d'arrêt (h)"]]
    for m in Maintenance.objects.select_related('echographe'):
        data.append([
            str(m.echographe), m.get_type_maintenance_display(),
            m.date_intervention.strftime('%d/%m/%Y'), m.technicien,
            (m.panne or '')[:40], (m.reparation or '')[:40], str(m.temps_arret),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B2E33')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7F6')]),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def export_maintenances_excel(request):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenances"
    entetes = ["Échographe", "Type", "Date", "Technicien", "Panne", "Réparation", "Temps d'arrêt (h)"]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for m in Maintenance.objects.select_related('echographe'):
        ws.append([
            str(m.echographe), m.get_type_maintenance_display(),
            m.date_intervention.strftime('%d/%m/%Y'), m.technicien,
            m.panne, m.reparation, m.temps_arret,
        ])

    for col in ws.columns:
        largeur = max(len(str(c.value)) if c.value else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(largeur, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="maintenances.xlsx"'
    wb.save(response)
    return response


def export_pannes_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pannes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Rapport des pannes — GMAO Biomédicale", styles['Title'])]

    data = [["Équipement", "Description", "Priorité", "Statut", "Technicien", "Déclarée le"]]
    for p in Panne.objects.select_related('equipement', 'technicien'):
        data.append([
            str(p.equipement), (p.description or '')[:50], p.get_priorite_display(),
            p.get_statut_display(),
            p.technicien.get_full_name() if p.technicien else "Non assigné",
            p.date_declaration.strftime('%d/%m/%Y %H:%M'),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B2E33')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7F6')]),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def export_pannes_excel(request):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pannes"
    entetes = ["Équipement", "Description", "Priorité", "Statut", "Technicien", "Déclarée le"]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in Panne.objects.select_related('equipement', 'technicien'):
        ws.append([
            str(p.equipement), p.description, p.get_priorite_display(), p.get_statut_display(),
            p.technicien.get_full_name() if p.technicien else "Non assigné",
            p.date_declaration.strftime('%d/%m/%Y %H:%M'),
        ])

    for col in ws.columns:
        largeur = max(len(str(c.value)) if c.value else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(largeur, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pannes.xlsx"'
    wb.save(response)
    return response
